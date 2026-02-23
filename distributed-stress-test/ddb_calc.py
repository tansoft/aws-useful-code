#!/usr/bin/env python3
"""
DynamoDB 费用精确计算工具 - 支持多列和多行模式

使用方法：
    python ddb_calc.py traffic1.json [traffic2.json ...]
    e.g.:
        python3 ddb_calc.py traffic_write.json traffic_read.json

注意事项：
    在生成每天的流量图时，会简化逻辑，直接按24小时，每小时的流量进行计算，忽略duration，repeat，samples等的计算
"""

import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.styles.borders import BORDER_THIN, BORDER_THICK, BORDER_MEDIUM

# DynamoDB 定价 (us-east-1 区域)
PRICING = {
    'on_demand_write': 0.625/1000000,
    'on_demand_read': 0.125/1000000,
    'provisioned_wcu_hour': 0.00065,
    'provisioned_rcu_hour': 0.00013,
    'reserved_wcu_hour_1y': 0.0128/100,
    'reserved_rcu_hour_1y': 0.0025/100,
    'reserved_wcu_hour_3y': 0.0081/100,
    'reserved_rcu_hour_3y': 0.0016/100,
    'reserved_upfront_wcu_100_1y': 150/100,
    'reserved_upfront_rcu_100_1y': 30/100,
    'reserved_upfront_wcu_100_3y': 180/100,
    'reserved_upfront_rcu_100_3y': 36/100,
    'storage_gb_month': 0.25,
}

def smooth_qpss(qpss):
    result = list(qpss)
    while True:
        changed = False
        for i in range(len(result)):
            if result[i] == 0:
                # 寻找左边第一个非零元素
                left = i - 1
                while left >= 0 and result[left] == 0:
                    left -= 1
                # 寻找右边第一个非零元素
                right = i + 1
                while right < len(result) and result[right] == 0:
                    right += 1
                
                left_val = result[left] if left >= 0 else 0.0
                right_val = result[right] if right < len(result) else 0.0
                
                if left_val > 0 or right_val > 0:
                    result[i] = (left_val + right_val) / 2
                    changed = True
        if not changed:
            break
    return result

def format_range(ws, cell_range, style_dict=None, **kwargs):
    """
    格式化指定范围的单元格
    
    Args:
        ws: worksheet 对象
        cell_range: 单元格范围，如 "A1:C5" 或 [(1,1), (5,3)] 或单个单元格 "A1"
        style_dict: 样式字典
        **kwargs: 格式化参数
    
    Example:
        # 方式1：使用关键字参数
        format_range(ws, "A1:C5",
                    font_name="Arial", 
                    font_size=12, 
                    bold=True,
                    bg_color="FFFF00",
                    horizontal="center",
                    number_format='#,##0.00')
        
        # 方式2：使用样式字典
        style = {
            'font': {'name': 'Arial', 'size': 12, 'bold': True, 'color': 'FF0000'},
            'fill': {'color': 'FFFF00'},
            'alignment': {'horizontal': 'center', 'vertical': 'middle'},
            'border': 'thin'
        }
        format_range(ws, "A1:C5", style_dict=style)

        # 格式化表头
        format_range(ws, cell_range, font_name="Arial", font_size=12, bold=True, bg_color="366092",
                font_color="FFFFFF", horizontal="center", vertical="middle", border="thin")
        # 格式化数据
        format_range(ws, cell_range, font_name="Arial", font_size=10, horizontal="left", vertical="middle", border="thin")
        # 格式化汇总
        format_range(ws, cell_range, font_name="Arial", font_size=11, bold=True, bg_color="D9D9D9", horizontal="right", vertical="middle", border="medium")
    """
    
    # 合并参数
    if style_dict:
        kwargs.update(style_dict)
    
    # 获取范围内的所有单元格
    try:
        cells = ws[cell_range]
        # 如果是单个单元格，转换为可迭代格式
        if hasattr(cells, 'coordinate'):
            cells = [[cells]]
        elif not isinstance(cells[0], (list, tuple)):
            cells = [cells]
    except:
        print(f"Invalid range: {cell_range}")
        return
    
    # 遍历所有单元格
    for row in cells:
        for cell in row:
            apply_cell_format(cell, **kwargs)

def apply_cell_format(cell, **kwargs):
    """应用格式到单个单元格"""
    
    # 字体设置
    font_params = {}
    if 'font' in kwargs:
        font_params.update(kwargs['font'])
    
    # 从直接参数中获取字体设置
    for key in ['font_name', 'font_size', 'bold', 'italic', 'font_color']:
        if key in kwargs:
            param_name = key.replace('font_', '') if key.startswith('font_') else key
            if param_name == 'font_color':
                param_name = 'color'
            font_params[param_name] = kwargs[key]
    
    if font_params:
        cell.font = Font(**font_params)
    
    # 背景填充
    if 'fill' in kwargs:
        fill_params = kwargs['fill']
        cell.fill = PatternFill(
            start_color=fill_params.get('color', 'FFFFFF'),
            end_color=fill_params.get('color', 'FFFFFF'),
            fill_type=fill_params.get('type', 'solid')
        )
    elif 'bg_color' in kwargs:
        cell.fill = PatternFill(
            start_color=kwargs['bg_color'],
            end_color=kwargs['bg_color'],
            fill_type='solid'
        )
    
    # 对齐方式
    alignment_params = {}
    if 'alignment' in kwargs:
        alignment_params.update(kwargs['alignment'])
    
    for key in ['horizontal', 'vertical', 'wrap_text']:
        if key in kwargs:
            alignment_params[key] = kwargs[key]
    
    if alignment_params:
        cell.alignment = Alignment(**alignment_params)
    
    # 边框设置
    if 'border' in kwargs:
        border_style = kwargs['border']
        if isinstance(border_style, str):
            side = Side(style=border_style)
            cell.border = Border(left=side, right=side, top=side, bottom=side)
        elif isinstance(border_style, dict):
            sides = {}
            for position in ['left', 'right', 'top', 'bottom']:
                if position in border_style:
                    sides[position] = Side(style=border_style[position])
                else:
                    sides[position] = Side(style=border_style.get('style', 'thin'))
            cell.border = Border(**sides)
    
    if 'number_format' in kwargs:
        cell.number_format = kwargs['number_format']

def parse_traffic(filenames):
    """解析traffic文件"""
    if isinstance(filenames, str):
        filenames = [filenames]
    
    stats = {'write_ops': [], 'read_ops': [], 'seed_columns': {}}
    
    def process_task(task):
        action = task.get('action', '')
        qps = task.get('qps', 0)
        times = task.get('times', 0)
        duration = task.get('duration', 0)
        samples = task.get('samples', 1)
        data = task.get('data', {})
        seed = task.get('seed', 0)
        seeds = task.get('seeds', [])
        qpss = task.get('qpss', [])
        
        # 记录seed的每一列数据
        if seed > 0 and action in ['putItem', 'updateItem', 'batchPutItem', 'deleteItem']:
            if seed not in stats['seed_columns']:
                stats['seed_columns'][seed] = []
            
            for col_name, col_value in data.items():
                # 解析列名和大小
                if isinstance(col_value, dict) and 'len' in col_value:
                    # randomkey_: {"r": 16, "len": 8000}
                    r = col_value.get('r', 1)
                    length = col_value['len']
                    actual_col_name = f"{col_name}{r}" if col_name.endswith('_') else col_name
                    stats['seed_columns'][seed].append({'name': actual_col_name, 'size_bytes': length})
                elif isinstance(col_value, int):
                    # 固定列名，数字表示长度
                    stats['seed_columns'][seed].append({'name': col_name, 'size_bytes': col_value})
                elif isinstance(col_value, str):
                    # 固定列名，字符串内容
                    stats['seed_columns'][seed].append({'name': col_name, 'size_bytes': len(col_value)})
        
        num_columns = len(data) if data else 1
        item_size = sum(v['len'] if isinstance(v, dict) and 'len' in v else (v if isinstance(v, int) else len(v) if isinstance(v, str) else 0) for v in data.values())
        item_size = max(item_size / 1024, 1)

        # 处理qpss+seeds
        if qpss and seeds:
            qpss = smooth_qpss(qpss)
            total_seeds = sum(seeds)
            for hour, qps_ratio in enumerate(qpss):
                if qps_ratio == 0:
                    continue
                hour_qps = qps * qps_ratio
                for seed_idx, seed_ratio in enumerate(seeds):
                    seed_num = seed_idx + 1
                    seed_qps = hour_qps * (seed_ratio / total_seeds)
                    
                    # 计算该seed的item大小和列数
                    seed_cols = stats['seed_columns'].get(seed_num, [])
                    seed_item_size = sum(c['size_bytes'] for c in seed_cols) / 1024 if seed_cols else item_size
                    seed_num_cols = len(seed_cols) if seed_cols else num_columns
                    
                    op_info = {
                        'action': action, 'qps': seed_qps, 'count': seed_qps * 3600,
                        'item_size_kb': seed_item_size, 'num_columns': seed_num_cols,
                        'hour': hour, 'seed': seed_num
                    }
                    
                    if action in ['getItem', 'getSubItem', 'batchGetItem', 'batchGetSubItem']:
                        stats['read_ops'].append(op_info)
            return
        
        # 常规处理
        op_count = times if times > 0 else (qps * duration if duration > 0 else 0)
        op_info = {'action': action, 'qps': qps, 'count': op_count, 'item_size_kb': item_size, 'num_columns': num_columns, 'hour': -1, 'seed': seed}
        
        if action in ['putItem', 'updateItem', 'deleteItem', 'batchPutItem']:
            if action == 'batchPutItem':
                op_info['count'] = op_count * samples
            stats['write_ops'].append(op_info)
        elif action in ['getItem', 'getSubItem', 'batchGetItem', 'batchGetSubItem']:
            if action in ['batchGetItem', 'batchGetSubItem']:
                op_info['count'] = op_count * samples
            stats['read_ops'].append(op_info)
    
    def traverse(obj):
        if isinstance(obj, list):
            for item in obj:
                traverse(item)
        elif isinstance(obj, dict) and 'action' in obj:
            process_task(obj)
    
    for filename in filenames:
        with open(filename, 'r') as f:
            traverse(json.load(f))
    
    return stats

def create_excel(stats, output_file):
    """生成Excel报表"""
    wb = openpyxl.Workbook()
    
    ws_base = wb.active
    ws_base.title = "基础数据"
    ws_output = wb.create_sheet("预算信息")
    
    # 生成基础数据表
    rows = create_base_sheet(ws_base, stats)    
    # 生成计算表
    create_calc_sheet(ws_output, rows)
    
    wb.save(output_file)
    print(f"✓ Excel报表已生成: {output_file}")
    print(f"✓ 包含：基础数据、预算信息")

def create_base_sheet(ws, stats):
    """创建基础数据表"""
    hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ht = Font(bold=True, color="FFFFFF")
    sf = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    st = Font(bold=True)
    
    row = 1
    rows = {}
    
    # Seed列定义表
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "Seed列定义"
    ws[f'A{row}'].fill = hf
    ws[f'A{row}'].font = ht
    row += 1
    
    headers = ["Seed", "列名", "列大小(Bytes)", "列大小(KB)", "列WCU", "列RCU"]
    for i, h in enumerate(headers):
        ws.cell(row, i+1, h).fill = sf
        ws.cell(row, i+1).font = st
    row += 1
    
    col_def_start = row
    for seed in sorted(stats['seed_columns'].keys()):
        for col_info in stats['seed_columns'][seed]:
            ws[f'A{row}'] = seed
            ws[f'B{row}'] = col_info['name']
            ws[f'C{row}'] = col_info['size_bytes']
            ws[f'D{row}'] = f"=C{row}/1024"
            ws[f'E{row}'] = f"=ROUNDUP(C{row}/1024,0)"
            ws[f'F{row}'] = f"=ROUNDUP(C{row}/4096,0)*0.5"
            row += 1
    col_def_end = row - 1
    row += 2
    format_range(ws, f"A{col_def_start}:C{col_def_end}", bg_color="FFC000")
    format_range(ws, f"C{col_def_start}:C{col_def_end}", number_format='#,##')
    format_range(ws, f"D{col_def_start}:D{col_def_end}", number_format='#,##0.00')

    # 写入操作明细
    ws.merge_cells(f'A{row}:M{row}')
    ws[f'A{row}'] = "写入操作明细"
    ws[f'A{row}'].fill = hf
    ws[f'A{row}'].font = ht
    row += 1
    
    # "列数", "Item大小(KB)", 
    headers = ["操作", "Seed", "时长", "QPS", "操作次数", "列数", "Item大小(KB)", "数据量(GB)", "多列WRU/次", "多列点写总WRU", "多列批写总WRU", "多行WRU/次", "多行总WRU"]
    for i, h in enumerate(headers):
        ws.cell(row, i+1, h).fill = sf
        ws.cell(row, i+1).font = st
    row += 1
    
    write_start = row
    for op in stats['write_ops']:
        seed = op.get('seed', 0)
        ws[f'A{row}'] = op['action']
        ws[f'B{row}'] = seed
        ws[f'C{row}'] = op['count']/op['qps']/3600 # op.get('hour', -1)
        ws[f'D{row}'] = op['qps']
        ws[f'E{row}'] = op['count']
        ws[f'F{row}'] = f"=COUNTIF($A${col_def_start}:$A${col_def_end},{seed})" #列数
        ws[f'G{row}'] = f"=SUMIF($A${col_def_start}:$A${col_def_end},{seed},$D${col_def_start}:$D${col_def_end})" #item大小
        ws[f'H{row}'] = f"=E{row}*G{row}/1024/1024" # 数据量
        ws[f'I{row}'] = f"=ROUNDUP(G{row},0)"  # 多列WRU/次，整个item
        ws[f'J{row}'] = f"=E{row}*F{row}*I{row}" # 多列点写总WRU
        ws[f'K{row}'] = f"=E{row}*I{row}" # 多列批写总WRU
        ws[f'L{row}'] = f"=SUMIF($A${col_def_start}:$A${col_def_end},{seed},E${col_def_start}:E${col_def_end})"  # 多行WRU/次：每列单独再汇总
        ws[f'M{row}'] = f"=E{row}*L{row}" # 多行总WRU
        # ws[f'F{row}'] = op['num_columns']
        # ws[f'G{row}'] = op['item_size_kb']
        row += 1
    write_end = row - 1
    ws[f'A{row}'] = "写入汇总"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{write_start}:C{write_end})"
    ws[f'E{row}'] = f"=SUM(E{write_start}:E{write_end})"
    ws[f'H{row}'] = f"=SUM(H{write_start}:H{write_end})"
    ws[f'I{row}'] = f"=SUM(I{write_start}:I{write_end})"
    ws[f'J{row}'] = f"=SUM(J{write_start}:J{write_end})"
    ws[f'K{row}'] = f"=SUM(K{write_start}:K{write_end})"
    ws[f'L{row}'] = f"=SUM(L{write_start}:L{write_end})"
    ws[f'M{row}'] = f"=SUM(M{write_start}:M{write_end})"
    rows['数据量(GB)'] = f'$H${row}'
    rows['多列总WRU(点写)'] = f'$J${row}'
    rows['多列总WRU(批写)'] = f'$K${row}'
    rows['多行总WRU'] = f'$M${row}'
    format_range(ws, f"A{write_start}:E{write_end}", bg_color="FFC000")
    format_range(ws, f"C{write_start}:C{write_end+1}", number_format='#,##0.00')
    format_range(ws, f"D{write_start}:E{write_end+1}", number_format='#,##')
    format_range(ws, f"G{write_start}:H{write_end+1}", number_format='#,##0.00')
    format_range(ws, f"J{write_start}:K{write_end+1}", number_format='#,##')
    format_range(ws, f"M{write_start}:M{write_end+1}", number_format='#,##')
    row += 3
    
    # 读取操作明细
    ws.merge_cells(f'A{row}:M{row}')
    ws[f'A{row}'] = "读取操作明细"
    ws[f'A{row}'].fill = hf
    ws[f'A{row}'].font = ht
    row += 1
    
    headers = ["操作", "Seed", "小时", "QPS", "操作次数", "列数", "Item大小(KB)", "多列RRU/次", "多列总RRU", "多列预置RCU","多行RRU/次", "多行总RRU", "多行预置RCU"]
    for i, h in enumerate(headers):
        ws.cell(row, i+1, h).fill = sf
        ws.cell(row, i+1).font = st
    row += 1
    
    read_start = row
    for op in stats['read_ops']:
        seed = op.get('seed', 0)
        ws[f'A{row}'] = op['action']
        ws[f'B{row}'] = seed
        ws[f'C{row}'] = op.get('hour', -1) # op['count']/op['qps']/3600
        ws[f'D{row}'] = op['qps']
        ws[f'E{row}'] = op['count']
        ws[f'F{row}'] = f"=COUNTIF($A${col_def_start}:$A${col_def_end},{seed})" #列数
        ws[f'G{row}'] = f"=SUMIF($A${col_def_start}:$A${col_def_end},{seed},$D${col_def_start}:$D${col_def_end})" #item大小
        ws[f'H{row}'] = f"=ROUNDUP(G{row}/4,0)*0.5"  # 多列RRU/次
        ws[f'I{row}'] = f"=E{row}*H{row}" # 多列总RRU
        ws[f'J{row}'] = f"=D{row}*H{row}" # 多列预置RCU
        ws[f'K{row}'] = f"=ROUNDUP(G{row}/4,0)*0.5"  # 多行RRU/次，通过Query
        ws[f'L{row}'] = f"=E{row}*K{row}" # 多行总RRU
        ws[f'M{row}'] = f"=D{row}*K{row}" # 多行预置RCU
        # ws[f'F{row}'] = op['num_columns']
        # ws[f'G{row}'] = op['item_size_kb']
        row += 1
    read_end = row - 1
    format_range(ws, f"A{read_start}:E{read_end}", bg_color="FFC000")
    format_range(ws, f"D{read_start}:E{read_end}", number_format='#,##')
    format_range(ws, f"G{read_start}:G{read_end}", number_format='#,##0.00')
    format_range(ws, f"I{read_start}:J{read_end}", number_format='#,##')
    format_range(ws, f"L{read_start}:M{read_end}", number_format='#,##')

    ws[f'O{read_start-1}'] = "多列RCU曲线"
    ws[f'P{read_start-1}'] = "多行RCU曲线"
    for i in range(24):
        ws[f'O{read_start+i}'] = f"=SUMIF(C{read_start}:C{read_end},{i},J{read_start}:J{read_end})"
        ws[f'P{read_start+i}'] = f"=SUMIF(C{read_start}:C{read_end},{i},M{read_start}:M{read_end})"

    ws[f'N{read_start+25}'] = "汇总计算"
    ws[f'N{read_start+25}'].font = Font(bold=True)

    ws[f'N{read_start+26}'] = "总RRU"
    ws[f'O{read_start+26}'] = f"=SUM(I{read_start}:I{read_end})"
    ws[f'P{read_start+26}'] = f"=SUM(L{read_start}:L{read_end})"
    ws[f'N{read_start+27}'] = "总RCU"
    ws[f'O{read_start+27}'] = f"=SUM(O{read_start}:O{read_start+23})"
    ws[f'P{read_start+27}'] = f"=SUM(P{read_start}:P{read_start+23})"
    ws[f'N{read_start+28}'] = "最大RCU"
    ws[f'O{read_start+28}'] = f"=MAX(O{read_start}:O{read_start+23})"
    ws[f'P{read_start+28}'] = f"=MAX(P{read_start}:P{read_start+23})"

    # 预留比预置的转折点
    # 预置WCU每小时0.00065 * WCU * 小时数 * 365 / 12 = 3年预留WCU每小时0.000081 * WCU * 24 * 365 / 12 + 3年预付1.8 * WCU / 36个月
    # 0.00065*WCU*hour*365/12=0.000081*WCU*24*365/12+1.80*WCU/36 => 5.5197471022
    reserve_hour = (PRICING['reserved_wcu_hour_3y']*24*365/12 + PRICING['reserved_upfront_wcu_100_3y']/36) / (PRICING['provisioned_wcu_hour'] * 365 / 12)
    ws[f'N{read_start+29}'] = "预留小时数"
    ws[f'O{read_start+29}'] = f"=ROUNDUP({reserve_hour},0)"
    ws[f'P{read_start+29}'] = f"=ROUNDUP({reserve_hour},0)"

    ws[f'N{read_start+30}'] = "预留RCU"
    ws[f'O{read_start+30}'] = f"=LARGE(O{read_start}:O{read_start+23},O{read_start+29})"
    ws[f'P{read_start+30}'] = f"=LARGE(P{read_start}:P{read_start+23},P{read_start+29})"

    ws[f'N{read_start+31}'] = "剩余预置RCU"
    ws[f'O{read_start+31}'] = f"=SUMIF(O{read_start}:O{read_start+23},\">\"&O{read_start+30},O{read_start}:O{read_start+23})-O{read_start+30}*COUNTIF(O{read_start}:O{read_start+23},\">\"&O{read_start+30})"
    ws[f'P{read_start+31}'] = f"=SUMIF(P{read_start}:P{read_start+23},\">\"&P{read_start+30},P{read_start}:P{read_start+23})-P{read_start+30}*COUNTIF(P{read_start}:P{read_start+23},\">\"&P{read_start+30})"

    format_range(ws, f"O{read_start}:P{read_start+31}", number_format='#,##')

    rows['多列总RRU'] = f'$O${read_start+26}'
    rows['多行总RRU'] = f'$P${read_start+26}'
    rows['多列总RCU'] = f'$O${read_start+27}'
    rows['多行总RCU'] = f'$P${read_start+27}'
    rows['多列最大RCU'] = f'$O${read_start+28}'
    rows['多行最大RCU'] = f'$P${read_start+28}'
    rows['预留小时数'] = f'$O${read_start+29}'

    rows['多列预留RCU'] = f'$O${read_start+30}'
    rows['多行预留RCU'] = f'$P${read_start+30}'
    # 购买预留后，没有覆盖的部分还需要叠加购买预置RCU
    rows['多列剩余预置RCU'] = f'$O${read_start+31}'
    rows['多行剩余预置RCU'] = f'$P${read_start+31}'

    # 格式化
    for i in range(ord('A'), ord('Q')):
        ws.column_dimensions[chr(i)].width = 18
    
    return rows

def create_calc_sheet(ws, rows):
    """创建计算表"""
    row = 1

    # 配置信息
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "配置信息"
    format_range(ws, f'A{row}', bg_color="366092", font_color="FFFFFF", bold=True)
    row += 1

    ws.append(['','按需','预置(小时)','预留(1年)','预留(3年)','预付(1年)','预付(3年)'])
    ws.append(['WRU/WCU',PRICING['on_demand_write'],PRICING['provisioned_wcu_hour'],
        PRICING['reserved_wcu_hour_1y'],PRICING['reserved_wcu_hour_3y'],
        PRICING['reserved_upfront_wcu_100_1y'],PRICING['reserved_upfront_wcu_100_3y']])
    ws.append(['RRU/RCU',PRICING['on_demand_read'],PRICING['provisioned_rcu_hour'],
        PRICING['reserved_rcu_hour_1y'],PRICING['reserved_rcu_hour_3y'],
        PRICING['reserved_upfront_rcu_100_1y'],PRICING['reserved_upfront_rcu_100_3y']])
    ws.append(['存储(GB/月)',PRICING['storage_gb_month'],'','最大WCU',5000000,'折扣off',0.0])
    base_row = row+1
    ws[f'F{base_row+3}'] = '预留购买小时数'
    ws[f'G{base_row+3}'] = f'=基础数据!{rows["预留小时数"]}'
    format_range(ws, f'E{base_row+2}', number_format='#,##', bg_color="FFC000")
    format_range(ws, f'G{base_row+2}', number_format='#,##0.00', bg_color="FFC000")
    row += 5

    # 基础数据
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "基础数据"
    format_range(ws, f'A{row}', bg_color="366092", font_color="FFFFFF", bold=True)
    row += 1

    stats_row = row + 1
    ws.append(['','多列点写','多列批写','多行模式'])
    ws.append(['数据量(GB)',f'=基础数据!{rows["数据量(GB)"]}',f'=基础数据!{rows["数据量(GB)"]}',f'=基础数据!{rows["数据量(GB)"]}'])
    ws.append(['总WRU',f'=基础数据!{rows["多列总WRU(点写)"]}',f'=基础数据!{rows["多列总WRU(批写)"]}',f'=基础数据!{rows["多行总WRU"]}'])
    ws.append(['写入小时',f'=ROUNDUP(B{stats_row+1}/$E${base_row+2}/3600,0)',f'=ROUNDUP(C{stats_row+1}/$E${base_row+2}/3600,0)',f'=ROUNDUP(D{stats_row+1}/$E${base_row+2}/3600,0)'])
    ws.append(['预留WCU',f'=IF(B{row+3}>=$G${base_row+3},$E${base_row+2},0)',f'=IF(C{row+3}>=$G${base_row+3},$E${base_row+2},0)',f'=IF(D{row+3}>=$G${base_row+3},$E${base_row+2},0)'])
    ws.append(['总RRU',f'=基础数据!{rows["多列总RRU"]}',f'=基础数据!{rows["多列总RRU"]}',f'=基础数据!{rows["多行总RRU"]}'])
    ws.append(['总RCU',f'=基础数据!{rows["多列总RCU"]}',f'=基础数据!{rows["多列总RCU"]}',f'=基础数据!{rows["多行总RCU"]}'])
    ws.append(['最大RCU',f'=基础数据!{rows["多列最大RCU"]}',f'=基础数据!{rows["多列最大RCU"]}',f'=基础数据!{rows["多行最大RCU"]}'])
    ws.append(['预留RCU',f'=基础数据!{rows["多列预留RCU"]}',f'=基础数据!{rows["多列预留RCU"]}',f'=基础数据!{rows["多行预留RCU"]}'])
    ws.append(['剩余预置RCU',f'=基础数据!{rows["多列剩余预置RCU"]}',f'=基础数据!{rows["多列剩余预置RCU"]}',f'=基础数据!{rows["多行剩余预置RCU"]}'])

    format_range(ws, f"B{stats_row}:D{stats_row}", number_format='#,##0.00')
    format_range(ws, f"B{stats_row+1}:D{stats_row+8}", number_format='#,##')

    row += 11

    # 费用汇总
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "按需费用汇总"
    format_range(ws, f'A{row}', bg_color="366092", font_color="FFFFFF", bold=True)
    row += 1

    ws[f'A{row}'] = "写入按需费用"
    ws[f'B{row}'] = f"=$B${base_row}*B{stats_row+1}*365/12"
    ws[f'C{row}'] = f"=$B${base_row}*C{stats_row+1}*365/12"
    ws[f'D{row}'] = f"=$B${base_row}*D{stats_row+1}*365/12"
    row += 1

    ws[f'A{row}'] = "读取按需费用"
    ws[f'B{row}'] = f"=$B${base_row+1}*B{stats_row+4}*365/12"
    ws[f'C{row}'] = f"=$B${base_row+1}*C{stats_row+4}*365/12"
    ws[f'D{row}'] = f"=$B${base_row+1}*D{stats_row+4}*365/12"
    row += 1

    ws[f'A{row}'] = "存储费用"
    ws[f'B{row}'] = f"=$B${base_row+2}*(B{stats_row}-25)"
    ws[f'C{row}'] = f"=$B${base_row+2}*(C{stats_row}-25)"
    ws[f'D{row}'] = f"=$B${base_row+2}*(D{stats_row}-25)"
    row += 1

    ws[f'A{row}'] = "总计"
    ws[f'B{row}'] = f"=sum(B{row-3}:B{row-1})"
    ws[f'C{row}'] = f"=sum(C{row-3}:C{row-1})"
    ws[f'D{row}'] = f"=sum(D{row-3}:D{row-1})"
    row += 1

    ws[f'A{row}'] = "折后总计"
    ws[f'B{row}'] = f"=B{row-1}*(1-$G${base_row+2})"
    ws[f'C{row}'] = f"=C{row-1}*(1-$G${base_row+2})"
    ws[f'D{row}'] = f"=D{row-1}*(1-$G${base_row+2})"
    format_range(ws, f"B{row-5}:D{row}", number_format='#,##0.00')
    row += 1

    row += 1

    # 预置费用汇总
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "预置费用汇总"
    format_range(ws, f'A{row}', bg_color="366092", font_color="FFFFFF", bold=True)
    row += 1

    ws[f'A{row}'] = "写入预置费用"
    ws[f'B{row}'] = f"=$C${base_row}*$E${base_row+2}*B{stats_row+2}*365/12"
    ws[f'C{row}'] = f"=$C${base_row}*$E${base_row+2}*C{stats_row+2}*365/12"
    ws[f'D{row}'] = f"=$C${base_row}*$E${base_row+2}*D{stats_row+2}*365/12"
    row += 1

    ws[f'A{row}'] = "读取预置费用"
    ws[f'B{row}'] = f"=$C${base_row+1}*B{stats_row+5}*365/12"
    ws[f'C{row}'] = f"=$C${base_row+1}*C{stats_row+5}*365/12"
    ws[f'D{row}'] = f"=$C${base_row+1}*D{stats_row+5}*365/12"
    row += 1

    ws[f'A{row}'] = "存储费用"
    ws[f'B{row}'] = f"=$B${base_row+2}*(B{stats_row}-25)"
    ws[f'C{row}'] = f"=$B${base_row+2}*(C{stats_row}-25)"
    ws[f'D{row}'] = f"=$B${base_row+2}*(D{stats_row}-25)"
    row += 1

    ws[f'A{row}'] = "总计"
    ws[f'B{row}'] = f"=sum(B{row-3}:B{row-1})"
    ws[f'C{row}'] = f"=sum(C{row-3}:C{row-1})"
    ws[f'D{row}'] = f"=sum(D{row-3}:D{row-1})"
    row += 1

    ws[f'A{row}'] = "折后总计"
    ws[f'B{row}'] = f"=B{row-1}*(1-$G${base_row+2})"
    ws[f'C{row}'] = f"=C{row-1}*(1-$G${base_row+2})"
    ws[f'D{row}'] = f"=D{row-1}*(1-$G${base_row+2})"
    format_range(ws, f"B{row-5}:D{row}", number_format='#,##0.00')
    row += 1

    row += 1

    # 预留费用汇总
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "预留费用汇总"
    format_range(ws, f'A{row}', bg_color="366092", font_color="FFFFFF", bold=True)
    row += 1

    ws[f'A{row}'] = "写入预留费用"
    ws[f'B{row}'] = f"=$E${base_row}*B{stats_row+3}*24*365/12+$G${base_row}*B{stats_row+3}/36"
    ws[f'C{row}'] = f"=$E${base_row}*C{stats_row+3}*24*365/12+$G${base_row}*C{stats_row+3}/36"
    ws[f'D{row}'] = f"=$E${base_row}*D{stats_row+3}*24*365/12+$G${base_row}*D{stats_row+3}/36"
    row += 1

    ws[f'A{row}'] = "写入预置费用"
    ws[f'B{row}'] = f"=$C${base_row}*($E${base_row+2}-B{stats_row+3})*B{stats_row+2}*365/12"
    ws[f'C{row}'] = f"=$C${base_row}*($E${base_row+2}-C{stats_row+3})*C{stats_row+2}*365/12"
    ws[f'D{row}'] = f"=$C${base_row}*($E${base_row+2}-D{stats_row+3})*D{stats_row+2}*365/12"
    row += 1

    ws[f'A{row}'] = "读取预留费用"
    ws[f'B{row}'] = f"=$E${base_row+1}*B{stats_row+7}*24*365/12+$G${base_row+1}*B{stats_row+7}/36"
    ws[f'C{row}'] = f"=$E${base_row+1}*C{stats_row+7}*24*365/12+$G${base_row+1}*C{stats_row+7}/36"
    ws[f'D{row}'] = f"=$E${base_row+1}*D{stats_row+7}*24*365/12+$G${base_row+1}*D{stats_row+7}/36"
    row += 1

    ws[f'A{row}'] = "读取预置费用"
    ws[f'B{row}'] = f"=$C${base_row+1}*B{stats_row+8}*365/12"
    ws[f'C{row}'] = f"=$C${base_row+1}*C{stats_row+8}*365/12"
    ws[f'D{row}'] = f"=$C${base_row+1}*D{stats_row+8}*365/12"
    row += 1

    ws[f'A{row}'] = "存储费用"
    ws[f'B{row}'] = f"=$B${base_row+2}*(B{stats_row}-25)"
    ws[f'C{row}'] = f"=$B${base_row+2}*(C{stats_row}-25)"
    ws[f'D{row}'] = f"=$B${base_row+2}*(D{stats_row}-25)"
    row += 1

    ws[f'A{row}'] = "总计"
    ws[f'B{row}'] = f"=sum(B{row-5}:B{row-1})"
    ws[f'C{row}'] = f"=sum(C{row-5}:C{row-1})"
    ws[f'D{row}'] = f"=sum(D{row-5}:D{row-1})"
    row += 1

    ws[f'A{row}'] = "折后总计"
    ws[f'B{row}'] = f"=B{row-1}*(1-$G${base_row+2})"
    ws[f'C{row}'] = f"=C{row-1}*(1-$G${base_row+2})"
    ws[f'D{row}'] = f"=D{row-1}*(1-$G${base_row+2})"
    format_range(ws, f"B{row-7}:D{row}", number_format='#,##0.00')
    row += 1

    # 格式化
    for i in range(ord('A'), ord('J')):
        ws.column_dimensions[chr(i)].width = 18

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python ddb_calc.py <traffic1.json> [traffic2.json ...]")
        sys.exit(1)
    
    traffic_files = sys.argv[1:]
    output_file = traffic_files[0].replace('.json', '.xlsx')
    
    print(f"正在解析 {len(traffic_files)} 个文件...")
    stats = parse_traffic(traffic_files)
    
    print(f"生成报表...")
    create_excel(stats, output_file)
