#!/usr/bin/env python3
import boto3
import yaml
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import json

def load_config(config_file):
    with open(config_file, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def get_lightsail_instances(region, config):
    client = boto3.client('lightsail', region_name=region)
    results = []
    try:
        bundles = client.get_bundles()['bundles']
        for bundle in bundles:
            if bundle.get('isActive'):
                # 检查操作系统平台
                supported_platforms = bundle.get('supportedPlatforms', [])
                os_config = config.get('operating_system', 'Linux')
                
                if os_config == 'Windows':
                    if 'WINDOWS' not in supported_platforms:
                        continue
                else:  # Linux, RHEL, SUSE都归为Linux平台
                    if 'LINUX_UNIX' not in supported_platforms:
                        continue
                
                cpu = bundle.get('cpuCount', 0)
                memory = bundle.get('ramSizeInGb', 0)
                storage = bundle.get('diskSizeInGb', 0)
                
                if config.get('cpu_min') and cpu < config['cpu_min']:
                    continue
                if config.get('cpu_max') and cpu > config['cpu_max']:
                    continue
                if config.get('memory_min') and memory < config['memory_min']:
                    continue
                if config.get('memory_max') and memory > config['memory_max']:
                    continue
                if config.get('storage', {}).get('size_min') and storage < config['storage']['size_min']:
                    continue
                if config.get('storage', {}).get('size_max') and storage > config['storage']['size_max']:
                    continue
                
                result = {
                    'service': 'Lightsail',
                    'region': region,
                    'instance_type': bundle['bundleId'],
                    'cpu': cpu,
                    'memory_gb': memory,
                    'storage_gb': storage,
                    'network_gbps': f"{bundle.get('transferPerMonthInGb', 0)}GB/月",
                    'price_monthly': bundle.get('price', 0),
                    'price_hourly': bundle.get('price', 0) / 730
                }
                
                if config.get('public_ip'):
                    result['public_ip_count'] = 1
                
                results.append(result)
    except Exception as e:
        print(f"Error fetching Lightsail in {region}: {e}")
    return results

def get_ec2_instances(region, config):
    client = boto3.client('ec2', region_name=region)
    results = []
    
    try:
        # 根据prefix过滤实例类型
        filters = []
        if config.get('instance_type_prefix'):
            prefix_list = list(config['instance_type_prefix'])
            instance_types = []
            for prefix in prefix_list:
                response = client.describe_instance_types(
                    Filters=[{'Name': 'instance-type', 'Values': [f'{prefix}*']}]
                )
                instance_types.extend(response['InstanceTypes'])
                while 'NextToken' in response:
                    response = client.describe_instance_types(
                        Filters=[{'Name': 'instance-type', 'Values': [f'{prefix}*']}],
                        NextToken=response['NextToken']
                    )
                    instance_types.extend(response['InstanceTypes'])
        else:
            response = client.describe_instance_types()
            instance_types = response['InstanceTypes']
            
            while 'NextToken' in response:
                response = client.describe_instance_types(NextToken=response['NextToken'])
                instance_types.extend(response['InstanceTypes'])
        
        for inst in instance_types:
            instance_type = inst['InstanceType']
            
            # 根据postfix过滤
            if config.get('instance_type_postfix'):
                postfix_config = config['instance_type_postfix']
                # 提取实例类型的后缀部分（如 m5.xlarge 中的基础部分）
                type_parts = instance_type.split('.')
                if len(type_parts) >= 2:
                    family = type_parts[0]  # 如 m5, m5d, m5n, m5dn, m5i, m5a, m5g, m7i-flex
                    # 提取基础家族名（数字前的部分+数字）
                    base_family = ''
                    for i, c in enumerate(family):
                        base_family += c
                        if c.isdigit():
                            break
                    
                    suffix = family[len(base_family):]  # 如 d, n, dn, i, a, g, i-flex
                    
                    if postfix_config == '-':
                        # 只要基本型：无后缀或只有处理器类型后缀(i/a/g)或flex变体
                        basic_suffixes = ['', 'i', 'a', 'g', 'i-flex', 'a-flex', 'g-flex']
                        if suffix not in basic_suffixes:
                            continue
                        
                        # 过滤掉非Nitro的老机型（代数小于5的，除了t3/t4）
                        generation = int(''.join(c for c in base_family if c.isdigit()))
                        family_prefix = ''.join(c for c in base_family if c.isalpha())
                        if generation < 5 and not family_prefix.startswith('t'):
                            continue
                    else:
                        # 检查是否包含指定的增强后缀
                        # 排除处理器类型后缀和flex
                        enhanced_suffix = suffix.replace('i', '').replace('a', '').replace('g', '').replace('-flex', '')
                        if enhanced_suffix:
                            # 有增强后缀，检查是否在允许列表中
                            if not any(s in enhanced_suffix for s in postfix_config):
                                continue
                        else:
                            # 没有增强后缀，跳过
                            continue
            
            cpu = inst['VCpuInfo']['DefaultVCpus']
            memory = inst['MemoryInfo']['SizeInMiB'] / 1024
            
            if config.get('cpu_min') and cpu < config['cpu_min']:
                continue
            if config.get('cpu_max') and cpu > config['cpu_max']:
                continue
            if config.get('memory_min') and memory < config['memory_min']:
                continue
            if config.get('memory_max') and memory > config['memory_max']:
                continue
            
            arch = inst['ProcessorInfo'].get('SupportedArchitectures', [])
            if config.get('architecture'):
                arch_list = config['architecture'] if isinstance(config['architecture'], list) else [config['architecture']]
                arch_map = {'x86': 'x86_64', 'arm': 'arm64', 'mac': 'x86_64_mac'}
                
                matched = False
                for arch_config in arch_list:
                    if arch_config == 'mac':
                        if inst['InstanceType'].startswith('mac'):
                            matched = True
                            break
                    elif arch_map.get(arch_config) in arch:
                        matched = True
                        break
                
                if not matched:
                    continue
            
            prices = get_ec2_prices(inst['InstanceType'], region, config.get('operating_system', 'Linux'), 
                                   config.get('pricing_types', {'ondemand': True}))
            
            result = {
                'service': 'EC2',
                'region': region,
                'instance_type': inst['InstanceType'],
                'cpu': cpu,
                'memory_gb': memory,
                'storage_gb': inst.get('InstanceStorageInfo', {}).get('TotalSizeInGB', 0),
                'network_gbps': inst.get('NetworkInfo', {}).get('NetworkPerformance', 'N/A')
            }
            
            result.update(prices)
            
            if config.get('public_ip'):
                network_info = inst.get('NetworkInfo', {})
                max_interfaces = network_info.get('MaximumNetworkInterfaces', 0)
                ipv4_per_interface = network_info.get('Ipv4AddressesPerInterface', 0)
                result['public_ip_count'] = max_interfaces * ipv4_per_interface
            
            results.append(result)
    except Exception as e:
        print(f"Error fetching EC2 in {region}: {e}")
    
    return results

def get_ec2_prices(instance_type, region, operating_system='Linux', pricing_types=None):
    if pricing_types is None:
        pricing_types = {'ondemand': True}
    prices = {}
    
    # 操作系统映射
    os_map = {
        'Linux': 'Linux',
        'Windows': 'Windows',
        'RHEL': 'RHEL',
        'SUSE': 'SUSE'
    }
    os_value = os_map.get(operating_system, 'Linux')
    
    # 区域名称映射
    region_map = {
        'us-east-1': 'US East (N. Virginia)',
        'us-east-2': 'US East (Ohio)',
        'us-west-1': 'US West (N. California)',
        'us-west-2': 'US West (Oregon)',
        'ap-south-1': 'Asia Pacific (Mumbai)',
        'ap-northeast-1': 'Asia Pacific (Tokyo)',
        'ap-northeast-2': 'Asia Pacific (Seoul)',
        'ap-northeast-3': 'Asia Pacific (Osaka)',
        'ap-southeast-1': 'Asia Pacific (Singapore)',
        'ap-southeast-2': 'Asia Pacific (Sydney)',
        'ca-central-1': 'Canada (Central)',
        'eu-central-1': 'Europe (Frankfurt)',
        'eu-west-1': 'Europe (Ireland)',
        'eu-west-2': 'Europe (London)',
        'eu-west-3': 'Europe (Paris)',
        'eu-north-1': 'Europe (Stockholm)',
        'sa-east-1': 'South America (Sao Paulo)'
    }
    
    location = region_map.get(region, region)
    
    # 获取按需价格和RI价格
    if pricing_types.get('ondemand') or any(k.startswith('ri_') for k in pricing_types.keys()):
        try:
            pricing_client = boto3.client('pricing', region_name='us-east-1')
            response = pricing_client.get_products(
            ServiceCode='AmazonEC2',
            Filters=[
                {'Type': 'TERM_MATCH', 'Field': 'instanceType', 'Value': instance_type},
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
                {'Type': 'TERM_MATCH', 'Field': 'operatingSystem', 'Value': os_value},
                {'Type': 'TERM_MATCH', 'Field': 'tenancy', 'Value': 'Shared'},
                {'Type': 'TERM_MATCH', 'Field': 'capacitystatus', 'Value': 'Used'},
                {'Type': 'TERM_MATCH', 'Field': 'preInstalledSw', 'Value': 'NA'}
            ],
            MaxResults=1
            )
            
            if response['PriceList']:
                price_data = json.loads(response['PriceList'][0])
            
            # 按需价格
            if pricing_types.get('ondemand') and 'OnDemand' in price_data['terms']:
                on_demand = list(price_data['terms']['OnDemand'].values())[0]
                price_dim = list(on_demand['priceDimensions'].values())[0]
                prices['ondemand'] = float(price_dim['pricePerUnit']['USD'])
            
            # RI价格
            if 'Reserved' in price_data['terms']:
                for term_key, term_val in price_data['terms']['Reserved'].items():
                    attrs = term_val['termAttributes']
                    if attrs['OfferingClass'] == 'standard':
                        lease_term = attrs['LeaseContractLength']
                        purchase_option = attrs['PurchaseOption']
                        price_dim = list(term_val['priceDimensions'].values())[0]
                        hourly_price = float(price_dim['pricePerUnit']['USD'])
                        
                        key_map = {
                            ('1yr', 'No Upfront'): 'ri_1y_no_upfront',
                            ('1yr', 'Partial Upfront'): 'ri_1y_partial_upfront',
                            ('1yr', 'All Upfront'): 'ri_1y_all_upfront',
                            ('3yr', 'No Upfront'): 'ri_3y_no_upfront',
                            ('3yr', 'Partial Upfront'): 'ri_3y_partial_upfront',
                            ('3yr', 'All Upfront'): 'ri_3y_all_upfront'
                        }
                        
                        price_key = key_map.get((lease_term, purchase_option))
                        if price_key and pricing_types.get(price_key):
                            prices[price_key] = hourly_price
        except Exception as e:
            print(f"  Warning: Pricing API failed for {instance_type}: {e}")
    
    # 获取Savings Plan价格
    if any(k.startswith('sp_') for k in pricing_types.keys()):
        try:
            savingsplans_client = boto3.client('savingsplans', region_name='us-east-1')
            
            # Compute Savings Plan
            if pricing_types.get('sp_compute_1y') or pricing_types.get('sp_compute_3y'):
                try:
                    offerings_response = savingsplans_client.describe_savings_plans_offerings(
                        planTypes=['Compute'],
                        paymentOptions=['No Upfront']
                    )
                    
                    for offering in offerings_response.get('searchResults', []):
                        duration = offering.get('durationSeconds')
                        offering_id = offering.get('offeringId')
                        
                        if offering_id:
                            rates_response = savingsplans_client.describe_savings_plans_offering_rates(
                                savingsPlanOfferingIds=[offering_id],
                                filters=[
                                    {'name': 'region', 'values': [region]},
                                    {'name': 'instanceType', 'values': [instance_type]},
                                    {'name': 'productDescription', 'values': [os_value]}
                                ]
                            )
                            
                            for rate in rates_response.get('searchResults', []):
                                if duration == 31536000 and pricing_types.get('sp_compute_1y'):
                                    prices['sp_compute_1y'] = float(rate.get('rate', 0))
                                elif duration == 94608000 and pricing_types.get('sp_compute_3y'):
                                    prices['sp_compute_3y'] = float(rate.get('rate', 0))
                except Exception as e:
                    print(f"  Warning: Compute SP failed for {instance_type}: {e}")
            
            # Instance Savings Plan
            if pricing_types.get('sp_instance_1y') or pricing_types.get('sp_instance_3y'):
                try:
                    offerings_response = savingsplans_client.describe_savings_plans_offerings(
                        planTypes=['EC2Instance'],
                        paymentOptions=['No Upfront']
                    )
                    
                    for offering in offerings_response.get('searchResults', []):
                        duration = offering.get('durationSeconds')
                        offering_id = offering.get('offeringId')
                        
                        if offering_id:
                            rates_response = savingsplans_client.describe_savings_plans_offering_rates(
                                savingsPlanOfferingIds=[offering_id],
                                filters=[
                                    {'name': 'region', 'values': [region]},
                                    {'name': 'instanceType', 'values': [instance_type]},
                                    {'name': 'productDescription', 'values': [os_value]}
                                ]
                            )
                            
                            for rate in rates_response.get('searchResults', []):
                                if duration == 31536000 and pricing_types.get('sp_instance_1y'):
                                    prices['sp_instance_1y'] = float(rate.get('rate', 0))
                                elif duration == 94608000 and pricing_types.get('sp_instance_3y'):
                                    prices['sp_instance_3y'] = float(rate.get('rate', 0))
                except Exception as e:
                    print(f"  Warning: Instance SP failed for {instance_type}: {e}")
        except Exception as e:
            print(f"  Warning: Savings Plans client failed: {e}")
    
    # 获取Spot价格（最近7天多AZ平均）
    if pricing_types.get('spot'):
        try:
            ec2_client = boto3.client('ec2', region_name=region)
            from datetime import datetime, timedelta
            
            spot_os_map = {
                'Linux': 'Linux/UNIX',
                'Windows': 'Windows',
                'RHEL': 'Red Hat Enterprise Linux',
                'SUSE': 'SUSE Linux'
            }
            spot_product = spot_os_map.get(operating_system, 'Linux/UNIX')
            
            start_time = datetime.utcnow() - timedelta(days=7)
            response = ec2_client.describe_spot_price_history(
                InstanceTypes=[instance_type],
                ProductDescriptions=[spot_product],
                StartTime=start_time
            )
            
            if response['SpotPriceHistory']:
                spot_prices = [float(item['SpotPrice']) for item in response['SpotPriceHistory']]
                prices['spot'] = sum(spot_prices) / len(spot_prices)
        except Exception as e:
            print(f"  Warning: Spot price failed for {instance_type}: {e}")
    
    return prices

def get_lambda_info(region, config):
    results = []
    memory_sizes = [128, 256, 512, 1024, 2048, 3072, 4096, 8192, 10240]
    
    for memory in memory_sizes:
        if config.get('memory_min') and memory / 1024 < config['memory_min']:
            continue
        if config.get('memory_max') and memory / 1024 > config['memory_max']:
            continue
        
        cpu = memory / 1769
        price_per_ms = 0.0000166667 * (memory / 1024)
        
        result = {
            'service': 'Lambda',
            'region': region,
            'instance_type': f'{memory}MB',
            'cpu': round(cpu, 2),
            'memory_gb': memory / 1024,
            'storage_gb': 512,
            'network_gbps': 'N/A',
            'price_hourly': price_per_ms * 3600000,
            'price_monthly': price_per_ms * 3600000 * 730
        }
        
        if config.get('public_ip'):
            result['public_ip_count'] = 0
        
        results.append(result)
    
    return results

def get_dto_price(region):
    """获取指定区域的DTO(Data Transfer Out)价格"""
    try:
        pricing_client = boto3.client('pricing', region_name='us-east-1')
        
        region_map = {
            'us-east-1': 'US East (N. Virginia)',
            'us-east-2': 'US East (Ohio)',
            'us-west-1': 'US West (N. California)',
            'us-west-2': 'US West (Oregon)',
            'ap-south-1': 'Asia Pacific (Mumbai)',
            'ap-northeast-1': 'Asia Pacific (Tokyo)',
            'ap-northeast-2': 'Asia Pacific (Seoul)',
            'ap-northeast-3': 'Asia Pacific (Osaka)',
            'ap-southeast-1': 'Asia Pacific (Singapore)',
            'ap-southeast-2': 'Asia Pacific (Sydney)',
            'ca-central-1': 'Canada (Central)',
            'eu-central-1': 'Europe (Frankfurt)',
            'eu-west-1': 'Europe (Ireland)',
            'eu-west-2': 'Europe (London)',
            'eu-west-3': 'Europe (Paris)',
            'eu-north-1': 'Europe (Stockholm)',
            'sa-east-1': 'South America (Sao Paulo)'
        }
        
        location = region_map.get(region, region)
        
        response = pricing_client.get_products(
            ServiceCode='AmazonEC2',
            Filters=[
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
                {'Type': 'TERM_MATCH', 'Field': 'transferType', 'Value': 'AWS Outbound'},
                {'Type': 'TERM_MATCH', 'Field': 'toLocation', 'Value': 'External'}
            ],
            MaxResults=1
        )
        
        if response['PriceList']:
            price_data = json.loads(response['PriceList'][0])
            on_demand = list(price_data['terms']['OnDemand'].values())[0]
            
            # 获取不同阶梯的价格
            tiers = []
            for price_dim in on_demand['priceDimensions'].values():
                begin_range = price_dim.get('beginRange', '0')
                end_range = price_dim.get('endRange', 'Inf')
                price_per_gb = float(price_dim['pricePerUnit']['USD'])
                tiers.append(f"{begin_range}-{end_range}GB: ${price_per_gb}/GB")
            
            return '; '.join(tiers) if tiers else '0'
    except Exception as e:
        print(f"  Warning: DTO price query failed for {region}: {e}")
        return 'N/A'

def generate_excel(results, output_file, config):
    wb = Workbook()
    wb.remove(wb.active)
    
    # 根据配置动态生成表头
    pricing_types = config.get('pricing_types', {'ondemand': True})
    
    base_headers = ['服务', '实例类型', 'CPU核心数', '内存(GB)', '存储(GB)', '网络性能']
    if config.get('public_ip'):
        base_headers.append('公网IP数量')
    
    hourly_headers = []
    monthly_headers = []
    price_keys = []
    
    if pricing_types.get('ondemand'):
        hourly_headers.append('按需/小时')
        monthly_headers.append('按需/月')
        price_keys.append('ondemand')
    
    ri_labels = {
        'ri_1y_no_upfront': 'RI-1年NoUp',
        'ri_1y_partial_upfront': 'RI-1年PartUp',
        'ri_1y_all_upfront': 'RI-1年AllUp',
        'ri_3y_no_upfront': 'RI-3年NoUp',
        'ri_3y_partial_upfront': 'RI-3年PartUp',
        'ri_3y_all_upfront': 'RI-3年AllUp'
    }
    
    for key, label in ri_labels.items():
        if pricing_types.get(key):
            hourly_headers.append(f'{label}/小时')
            monthly_headers.append(f'{label}/月')
            price_keys.append(key)
    
    if pricing_types.get('sp_compute_1y'):
        hourly_headers.append('SP-Compute-1年/小时')
        monthly_headers.append('SP-Compute-1年/月')
        price_keys.append('sp_compute_1y')
    
    if pricing_types.get('sp_compute_3y'):
        hourly_headers.append('SP-Compute-3年/小时')
        monthly_headers.append('SP-Compute-3年/月')
        price_keys.append('sp_compute_3y')
    
    if pricing_types.get('sp_instance_1y'):
        hourly_headers.append('SP-Instance-1年/小时')
        monthly_headers.append('SP-Instance-1年/月')
        price_keys.append('sp_instance_1y')
    
    if pricing_types.get('sp_instance_3y'):
        hourly_headers.append('SP-Instance-3年/小时')
        monthly_headers.append('SP-Instance-3年/月')
        price_keys.append('sp_instance_3y')
    
    if pricing_types.get('spot'):
        hourly_headers.append('Spot均价/小时')
        monthly_headers.append('Spot均价/月')
        price_keys.append('spot')
    
    headers = base_headers + hourly_headers + monthly_headers
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # 按区域分组
    region_results = {}
    for result in results:
        region = result['region']
        if region not in region_results:
            region_results[region] = []
        region_results[region].append(result)
    
    # 为每个区域创建工作表
    for region, region_data in region_results.items():
        ws = wb.create_sheet(title=region)
        
        # 如果需要公网IP，添加DTO价格信息
        if config.get('public_ip'):
            dto_price = get_dto_price(region)
            ws.cell(row=1, column=1, value=f"DTO价格: {dto_price}")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            start_row = 3
        else:
            start_row = 1
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # 写入数据
        for row_idx, result in enumerate(region_data, start_row + 1):
            col = 1
            ws.cell(row=row_idx, column=col, value=result['service'])
            col += 1
            ws.cell(row=row_idx, column=col, value=result['instance_type'])
            col += 1
            ws.cell(row=row_idx, column=col, value=result['cpu'])
            col += 1
            ws.cell(row=row_idx, column=col, value=result['memory_gb'])
            col += 1
            ws.cell(row=row_idx, column=col, value=result['storage_gb'])
            col += 1
            ws.cell(row=row_idx, column=col, value=str(result['network_gbps']))
            col += 1
            
            if config.get('public_ip'):
                ws.cell(row=row_idx, column=col, value=result.get('public_ip_count', 0))
                col += 1
            
            # 输出小时价格
            hourly_col_start = col
            for price_key in price_keys:
                ws.cell(row=row_idx, column=col, value=result.get(price_key, 0))
                col += 1
            
            # 输出月度价格（使用公式）
            for i, price_key in enumerate(price_keys):
                col_letter = chr(64 + hourly_col_start + i) if hourly_col_start + i <= 26 else f"A{chr(64 + hourly_col_start + i - 26)}"
                ws.cell(row=row_idx, column=col, value=f"={col_letter}{row_idx}*730")
                col += 1
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 14
    
    wb.save(output_file)
    print(f"报告已生成: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='AWS机型选择程序')
    parser.add_argument('-c', '--config', default='config.yaml', help='配置文件路径')
    args = parser.parse_args()
    
    config = load_config(args.config)
    regions = config.get('regions', ['us-east-1'])
    services = config.get('services', ['lightsail', 'ec2', 'lambda'])
    
    all_results = []
    
    for region in regions:
        print(f"正在查询区域: {region}")
        
        if 'lightsail' in services:
            print(f"  - 查询 Lightsail...")
            all_results.extend(get_lightsail_instances(region, config))
        
        if 'ec2' in services:
            print(f"  - 查询 EC2...")
            all_results.extend(get_ec2_instances(region, config))
        
        if 'lambda' in services:
            print(f"  - 查询 Lambda...")
            all_results.extend(get_lambda_info(region, config))
    
    output_file = f"aws_instance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    generate_excel(all_results, output_file, config)
    print(f"\n共找到 {len(all_results)} 个符合条件的配置")

if __name__ == '__main__':
    main()
